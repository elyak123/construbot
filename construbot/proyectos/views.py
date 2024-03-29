import importlib
import json
from decimal import Decimal
from django import shortcuts
from django.conf import settings
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView, TemplateView
from django.urls import reverse, reverse_lazy
from django.db.models import Max, F, Q
from django.db.models.functions import Lower
from django.http import JsonResponse
from django.contrib.auth import get_user_model
from openpyxl import load_workbook
from construbot.users.models import Company, NivelAcceso
from construbot.proyectos import forms
from construbot.core.utils import BasicAutocomplete, get_object_403_or_404
from .apps import ProyectosConfig
from .models import Contrato, Contraparte, Sitio, Units, Concept, Destinatario, Estimate, Retenciones
from .utils import contratosvigentes, estimacionespendientes_facturacion, estimacionespendientes_pago,\
    totalsinfacturar, total_sinpago, path_processing

try:
    auth = importlib.import_module(settings.CONSTRUBOT_AUTHORIZATION_CLASS)
except ImportError:
    from construbot.users import auth

User = get_user_model()


class ProyectosMenuMixin(auth.AuthenticationTestMixin):
    permiso_requerido = 2
    app_label_name = ProyectosConfig.verbose_name
    menu_specific = [
        {
            'title': 'Contratos',
            'url': 'construbot.proyectos:listado_de_contratos',
            'always_appear': True,
            'icon': 'bookmark',
            'parent': False,
            'type': 'submenu',
            'submenu': [],
        },
        {
            'title': 'Clientes',
            'url': 'construbot.proyectos:listado_de_clientes',
            'always_appear': True,
            'icon': 'person',
            'parent': False,
            'type': 'submenu',
            'submenu': [],
        },
        {
            'title': 'Ubicaciones',
            'url': 'construbot.proyectos:listado_de_sitios',
            'always_appear': True,
            'icon': 'map-marker',
            'parent': False,
            'type': 'submenu',
            'submenu': [],
        },
        {
            'title': 'Contactos',
            'url': 'construbot.proyectos:listado_de_destinatarios',
            'always_appear': True,
            'icon': 'people',
            'parent': False,
            'type': 'submenu',
            'submenu': [],
        },
        {
            'title': 'Unidades',
            'url': 'construbot.proyectos:catalogo_de_unidades',
            'always_appear': True,
            'icon': 'eyedropper',
            'parent': False,
            'type': 'submenu',
            'submenu': [],
        },
    ]
    model_options = {
        'Contrato': {
            'model': Contrato,
        },
        'Contraparte': {
            'ordering': 'cliente_name',
            'model': Contraparte,
        },
        'Sitio': {
            'ordering': 'sitio_name',
            'model': Sitio,
        },
        'Destinatario': {
            'ordering': 'destinatario_text',
            'model': Destinatario,
        },
        'Estimate': {
            'model': Estimate,
        }
    }

    def get_company_query(self, opcion):
        company_query = {
            'Contrato': {
                'contraparte__company': self.request.user.currently_at,
            },
            'Contraparte': {
                'company': self.request.user.currently_at
            },
            'Sitio': {
                'cliente__company': self.request.user.currently_at
            },
            'Destinatario': {
                'contraparte__company': self.request.user.currently_at
            },
            'Estimate': {
                'project__contraparte__company': self.request.user.currently_at
            },
        }
        return company_query[opcion]


class ProyectDashboardView(ProyectosMenuMixin, ListView):
    permiso_requerido = 1
    template_name = 'proyectos/index.html'
    model = Contrato

    def get_context_data(self, **kwargs):
        self.object_list = self.get_queryset()
        context = super(ProyectDashboardView, self).get_context_data(**kwargs)
        context['object'] = self.request.user.currently_at
        context['c_object'] = contratosvigentes(self.request.user)
        context['estimacionespendientes_facturacion'] = estimacionespendientes_facturacion(
            self.request.user.currently_at, context['almenos_coordinador'], self.request.user
        )
        context['estimacionespendientes_pago'] = estimacionespendientes_pago(
            self.request.user.currently_at, context['almenos_coordinador'], self.request.user
        )
        context['total_sin_facturar'] = totalsinfacturar(context['estimacionespendientes_facturacion'])
        context['total_sinpago'] = total_sinpago(context['estimacionespendientes_pago'])
        return context


class DynamicList(ProyectosMenuMixin, ListView):
    paginate_by = 10

    def get_queryset(self):
        if self.request.user.nivel_acceso.nivel >= 3:
            self.queryset = self.model.objects.filter(
                **self.get_company_query(self.model.__name__)).order_by(
                Lower(self.model_options[self.model.__name__]['ordering'])
            )
        else:
            self.queryset = Contrato.especial.asignaciones(self.request.user, self.model).order_by(
                    Lower(self.model_options[self.model.__name__]['ordering']))
        return super(DynamicList, self).get_queryset()

    def get_context_data(self, **kwargs):
        context = super(DynamicList, self).get_context_data(**kwargs)
        context['model'] = self.model.__name__
        return context


class ContratoListView(DynamicList):
    model = Contrato
    ordering = '-fecha'

    def get_queryset(self):
        if self.request.user.nivel_acceso.nivel >= 3:
            self.queryset = self.model.objects.filter(
                contraparte__company=self.request.user.currently_at).order_by(self.ordering)
        else:
            self.queryset = self.model.objects.filter(
                users=self.request.user,
                **self.get_company_query(self.model.__name__)
            ).order_by(self.ordering)
        return self.queryset


class ClienteListView(DynamicList):
    model = Contraparte


class SitioListView(DynamicList):
    model = Sitio


class DestinatarioListView(DynamicList):
    model = Destinatario

    def get_queryset(self):
        if self.request.user.nivel_acceso.nivel < 3:
            clientes = Contrato.especial.asignaciones(self.request.user, Contraparte)
            self.queryset = Destinatario.objects.filter(contraparte__in=clientes).order_by(
                Lower(self.model_options[self.model.__name__]['ordering'])
            )
            return self.queryset
        return super(DestinatarioListView, self).get_queryset()


class CatalogoConceptos(ProyectosMenuMixin, ListView):
    model = Concept
    ordering = 'code'
    permiso_requerido = 3
    nivel_permiso_asignado = 2
    asignacion_requerida = True

    def get_assignment_args(self):
        self.contrato = self.get_contrato()
        return self.contrato, self.request.user.contrato_set.all()

    def get_contrato(self):
        self.contrato = get_object_403_or_404(
            Contrato, self.request.user, pk=self.kwargs['pk'], contraparte__company=self.request.user.currently_at
        )
        return self.contrato

    def get(self, request, *args, **kwargs):
        queryset = self.model.objects.filter(project=self.contrato).order_by('pk')
        json = {}
        json['conceptos'] = []
        for concepto in queryset:
            aux = {}
            aux["code"] = concepto.code
            aux["concept_text"] = concepto.concept_text
            aux["unit"] = concepto.unit.unit
            aux["cuantity"] = concepto.total_cuantity
            aux["unit_price"] = concepto.unit_price
            json['conceptos'].append(aux)
        return JsonResponse(json)


class DynamicDetail(ProyectosMenuMixin, DetailView):
    permiso_requerido = 3
    asignacion_requerida = True
    change_company_ability = False

    def get_assignment_args(self):
        self.object = self.get_object()
        return self.object, Contrato.especial.asignaciones(self.request.user, self.model)

    def get_context_object_name(self, obj):
        return obj.__class__.__name__.lower()

    def get_object(self, queryset=None):
        if not hasattr(self, 'object'):
            self.object = get_object_403_or_404(
                self.model, self.request.user, pk=self.kwargs['pk'], **self.get_company_query(self.model.__name__)
            )
        return self.object

    def contratos_ordenados(self):
        if self.request.user.nivel_acceso.nivel >= self.permiso_requerido:
            return self.object.get_contratos_ordenados()
        contratos = self.object.contrato_set.filter(
            users=self.request.user,
            **self.get_company_query('Contrato')).order_by('-fecha')
        return contratos

    def get_context_data(self, **kwargs):
        context = super(DynamicDetail, self).get_context_data(**kwargs)
        if self.model is Sitio or self.model is Contraparte:
            context['contratos_ordenados'] = self.contratos_ordenados()
        return context


class ContratoDetailView(DynamicDetail):
    permiso_requerido = 3
    asignacion_requerida = True
    model = Contrato

    def get_assignment_args(self):
        self.object = self.get_object()
        return self.object, self.request.user.contrato_set.all()

    def get_object(self, queryset=None):
        query_kw = self.get_company_query(self.model.__name__)
        query_kw.update({'pk': self.kwargs['pk']})
        return get_object_403_or_404(self.model, self.request.user, **query_kw)


class ClienteDetailView(DynamicDetail):
    model = Contraparte


class SitioDetailView(DynamicDetail):
    model = Sitio


class DestinatarioDetailView(DynamicDetail):
    # no hay info confidencial comprometida
    permiso_requerido = 1
    asignacion_requerida = False
    model = Destinatario


class EstimateDetailView(DynamicDetail):
    permiso_requerido = 3
    asignacion_requerida = True
    model = Estimate

    def get_assignment_args(self):
        self.object = self.get_object()
        return self.object.project, self.request.user.contrato_set.all()

    def get_context_data(self, **kwargs):
        context = super(EstimateDetailView, self).get_context_data(**kwargs)
        context["conceptos"] = self.object.anotaciones_conceptos()
        context["total_estimacion"] = context["conceptos"].importe_total_esta_estimacion()['total']
        context["cantidad_de_conceptos"] = len(context["conceptos"])
        return context


class SubcontratosReport(EstimateDetailView):
    template_name = 'proyectos/reporte_estimaciones_subcontratos.html'

    def get_context_data(self, *args, **kwargs):
        #  Llamamos al super del padre para evitar la ejecucion de queries que no necesitamos.
        context = super(EstimateDetailView, self).get_context_data(*args, **kwargs)
        path = path_processing(self.object.project.path)
        context['subestimaciones'] = Estimate.especial.reporte_subestimaciones(
            self.object.start_date, self.object.finish_date, self.object.project.depth, path)
        context['acumulado'] = Estimate.especial.total_acumulado_subestimaciones(
            self.object.start_date, self.object.finish_date, self.object.project.depth, path
        )[0]
        context['actual'] = Estimate.especial.total_actual_subestimaciones(
            self.object.start_date, self.object.finish_date, self.object.project.depth, path
        )[0]
        context['anterior'] = Estimate.especial.total_anterior_subestimaciones(
            self.object.start_date, self.object.finish_date, self.object.project.depth, path
        )[0]
        context['contratado'] = Estimate.especial.total_contratado_subestimaciones(
            self.object.start_date, self.object.finish_date, self.object.project.depth, path
        )[0]
        return context


class EstimatePdfPrint(EstimateDetailView):
    nivel_permiso_asignado = 2
    template_name = 'proyectos/concept_pdf_estimate.html'


class GeneratorPdfPrint(EstimateDetailView):
    template_name = 'proyectos/concept_pdf_generator.html'


class DummyFileForm(ProyectosMenuMixin, TemplateView):
    template_name = 'core/dummy_input.html'


class ContratoCreationView(ProyectosMenuMixin, CreateView):
    permiso_requerido = 2
    change_company_ability = False
    form_class = forms.ContratoForm
    template_name = 'proyectos/creation_form.html'

    def get_form(self, *args, **kwargs):
        form = super(ContratoCreationView, self).get_form(form_class=None)
        form.request = self.request
        return form

    def get_max_id(self):
        max_id = self.form_class.Meta.model.objects.filter(
            contraparte__company=self.request.user.currently_at
        ).aggregate(Max('folio'))['folio__max'] or 0
        return max_id

    def get_depth(self):
        return 1

    def get_initial(self):
        initial_obj = super(ContratoCreationView, self).get_initial()
        max_id = self.get_max_id()
        max_id += 1
        initial_obj['currently_at'] = self.request.user.currently_at.company_name
        initial_obj['folio'] = max_id
        initial_obj['users'] = [self.request.user.pk]
        initial_obj['depth'] = self.get_depth()
        initial_obj['path'] = 'random'
        initial_obj['numchild'] = 0
        return initial_obj

    def get_context_data(self, **kwargs):
        context = super(ContratoCreationView, self).get_context_data(**kwargs)
        context['company'] = self.request.user.currently_at
        # context['dummy_file'] = True
        return context


class SubcontratoCreationView(ContratoCreationView):
    form_class = forms.SubContratoForm

    def dispatch(self, request, *args, **kwargs):
        self.contrato = Contrato.objects.get(pk=self.kwargs['pk'])
        return super().dispatch(request, *args, **kwargs)

    def get_form(self, *args, **kwargs):
        form = super(SubcontratoCreationView, self).get_form(form_class=None)
        form.contrato = self.contrato
        return form

    def get_max_id(self):
        return self.contrato.get_children_count()

    def get_depth(self):
        return self.contrato.get_depth() + 1

    def get_initial(self):
        obj = super(SubcontratoCreationView, self).get_initial()
        obj.update({'sitio': self.contrato.sitio})
        return obj

    def get_context_data(self, **kwargs):
        context = super(ContratoCreationView, self).get_context_data(**kwargs)
        context['subcontrato'] = True
        return context


class DynamicCreation(ProyectosMenuMixin, CreateView):
    permiso_requerido = 1
    change_company_ability = False
    template_name = 'proyectos/creation_form.html'

    def get_initial(self):
        initial_obj = super(DynamicCreation, self).get_initial()
        initial_obj['company'] = self.request.user.currently_at
        return initial_obj

    def get_form(self, *args, **kwargs):
        form = super(DynamicCreation, self).get_form(form_class=None)
        form.request = self.request
        return form


class ClienteCreationView(DynamicCreation):
    form_class = forms.ClienteForm


class SitioCreationView(DynamicCreation):
    form_class = forms.SitioForm


class DestinatarioCreationView(DynamicCreation):
    form_class = forms.DestinatarioForm


class EstimateCreationView(ProyectosMenuMixin, CreateView):
    change_company_ability = False
    permiso_requerido = 1
    asignacion_requerida = True
    form_class = forms.EstimateForm
    model = Contrato
    template_name = 'proyectos/estimate_form.html'

    def get_assignment_args(self):
        self.project_instance = self.get_project_instance()
        return self.project_instance, self.request.user.contrato_set.all()

    def get_project_instance(self):
        if not hasattr(self, 'project_instance'):
            self.project_instance = get_object_403_or_404(
                Contrato,
                self.request.user,
                pk=self.kwargs.get('pk'),
                contraparte__company=self.request.user.currently_at
            )
        return self.project_instance

    def get(self, request, *args, **kwargs):
        self.object = None
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        fill_data = self.fill_concept_formset()
        inlineform = forms.estimateConceptInlineForm(count=self.concept_count)
        return self.render_to_response(
            self.get_context_data(**self.get_prepare_generator_context(form, inlineform, fill_data))
        )

    def post(self, request, *args, **kwargs):
        self.object = None
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        fill_data = self.fill_concept_formset()
        if form.is_valid():
            return self.form_valid(form, fill_data)
        else:
            inlineform = forms.estimateConceptInlineForm(count=self.concept_count)
            return self.form_invalid(form, inlineform, fill_data)

    def get_prepare_generator_context(self, form, inlineform, fill_data):
        if hasattr(inlineform, 'form') and callable(inlineform):
            if not hasattr(inlineform, 'is_bound') or not inlineform.is_bound:
                generator_inline_concept = inlineform(initial=fill_data)
        else:
            generator_inline_concept = inlineform
        codes = [x['concept'].code for x in fill_data]
        image_formset_prefix = [x.nested.prefix for x in generator_inline_concept.forms]
        vertice_formset_prefix = [x.vertices.prefix for x in generator_inline_concept.forms]
        autocomplete_data = [x['concept_text_input'] for x in fill_data]
        return_dict = {
            'form': form,
            'generator_inline_concept': generator_inline_concept,
            'generator_zip': zip(generator_inline_concept, codes),
            'generator_autocomplete': zip(autocomplete_data, codes),
            'image_formset_prefix': image_formset_prefix,
            'vertice_formset_prefix': vertice_formset_prefix
        }
        return return_dict

    def fill_concept_formset(self):
        concepts = Concept.objects.filter(project=self.get_project_instance()).order_by('id')
        data = [
            {
                'concept': x, 'cuantity_estimated': 0,
                'concept_text_input': x.concept_text
            } for x in concepts
        ]
        self.concept_count = concepts.count()
        return data

    def get_initial(self):
        initial_dict = super(EstimateCreationView, self).get_initial()

        max_consecutive = Estimate._meta.model._default_manager.filter(
            project=self.kwargs['pk']
        ).aggregate(Max('consecutive'))['consecutive__max'] or 0
        max_consecutive += 1

        initial_dict['consecutive'] = max_consecutive
        initial_dict['project'] = self.kwargs['pk']
        initial_dict['draft_by'] = self.request.user
        return initial_dict

    def form_valid(self, form, fill_data):
        inlineform = forms.estimateConceptInlineForm(count=self.concept_count)
        generator_inline_concept = inlineform(
            self.request.POST,
            self.request.FILES,
            instance=form.instance
        )
        if generator_inline_concept.is_valid():
            response = super(EstimateCreationView, self).form_valid(form)
            generator_inline_concept.save()
            return response
        else:
            return self.form_invalid(form, generator_inline_concept, fill_data)

    def form_invalid(self, form, inlineform, fill_data):
        return self.render_to_response(
            self.get_context_data(**self.get_prepare_generator_context(form, inlineform, fill_data))
        )

    def get_success_url(self):
        url = reverse_lazy('proyectos:contrato_detail', kwargs={
            'pk': self.kwargs['pk']
        })
        return url

    def get_context_data(self, **kwargs):
        context = super(EstimateCreationView, self).get_context_data(**kwargs)
        context['project_instance'] = self.project_instance
        return context


class DynamicEdition(ProyectosMenuMixin, UpdateView):
    permiso_requerido = 1
    change_company_ability = False
    template_name = 'proyectos/creation_form.html'

    def get_object(self):
        obj = get_object_403_or_404(
            self.form_class.Meta.model,
            self.request.user,
            pk=self.kwargs['pk'],
            **self.get_company_query(self.form_class.Meta.model.__name__)
        )
        return obj

    def get_initial(self):
        initial_obj = super(DynamicEdition, self).get_initial()
        initial_obj['company'] = self.request.user.currently_at
        return initial_obj

    def get_form(self, *args, **kwargs):
        form = super(DynamicEdition, self).get_form(form_class=None)
        form.request = self.request
        return form


class ContratoEditView(ProyectosMenuMixin, UpdateView):
    permiso_requerido = 2
    nivel_permiso_asignado = 2
    asignacion_requerida = True
    change_company_ability = False
    form_class = forms.ContratoForm
    template_name = 'proyectos/creation_form.html'

    def get_assignment_args(self):
        self.object = self.get_object()
        return self.object, self.request.user.contrato_set.all()

    def get_form(self, *args, **kwargs):
        form = super(ContratoEditView, self).get_form(form_class=None)
        form.request = self.request
        return form

    def get_object(self):
        if not hasattr(self, 'object'):
            self.object = get_object_403_or_404(
                Contrato,
                self.request.user,
                pk=self.kwargs['pk'],
                contraparte__company=self.request.user.currently_at
            )
        return self.object

    def get_context_data(self, **kwargs):
        context = super(ContratoEditView, self).get_context_data(**kwargs)
        # context['dummy_file'] = True
        return context

    def get_initial(self):
        initial_obj = super(ContratoEditView, self).get_initial()
        initial_obj['currently_at'] = self.request.user.currently_at.company_name
        return initial_obj


class ClienteEditView(DynamicEdition):
    form_class = forms.ClienteForm


class SitioEditView(DynamicEdition):
    form_class = forms.SitioForm


class DestinatarioEditView(DynamicEdition):
    form_class = forms.DestinatarioForm


class EstimateEditView(ProyectosMenuMixin, UpdateView):
    permiso_requerido = 3
    asignacion_requerida = True
    nivel_permiso_asignado = 1
    change_company_ability = False
    form_class = forms.EstimateForm
    template_name = 'proyectos/estimate_form.html'
    model = Estimate

    def get_assignment_args(self):
        self.object = self.get_object()
        return self.object.project, self.request.user.contrato_set.all()

    def get_object(self, queryset=None):
        if not hasattr(self, 'object'):
            self.object = get_object_403_or_404(
                self.model,
                self.request.user,
                pk=self.kwargs['pk'],
                project__contraparte__company=self.request.user.currently_at
            )
        return self.object

    def get_initial(self):
        init_dict = {}
        init_dict['draft_by'] = self.request.user
        init_dict['project'] = self.kwargs['pk']
        return super(EstimateEditView, self).get_initial()

    def form_valid(self, form):
        form.save()
        conceptformclass = forms.estimateConceptInlineForm()
        self.conceptForm = conceptformclass(
            self.request.POST,
            self.request.FILES,
            instance=self.object
        )
        if self.conceptForm.is_valid():
            self.conceptForm.save()
            return super(EstimateEditView, self).form_valid(form)
        else:
            return self.form_invalid(form)

    def get_formset_for_context(self):
        if not hasattr(self, 'conceptForm'):
            formset = forms.estimateConceptInlineForm()(instance=self.object)
        else:
            formset = self.conceptForm
        return formset

    def get_concept_codes(self):
        estimate_concepts = Concept.objects.filter(project=self.object.project).order_by('id')
        concepts = [x.code for x in estimate_concepts]
        self.estimate_concepts_names = [x.concept_text for x in estimate_concepts]
        return concepts

    def get_context_data(self, **kwargs):
        context = super(EstimateEditView, self).get_context_data(**kwargs)
        formset = self.get_formset_for_context()
        concept_codes = self.get_concept_codes()
        image_formset_prefix = [x.nested.prefix for x in formset.forms]
        context['generator_inline_concept'] = formset
        context['generator_zip'] = zip(formset, concept_codes)
        context['generator_autocomplete'] = zip(self.estimate_concepts_names, concept_codes)
        context['image_formset_prefix'] = image_formset_prefix
        context['project_instance'] = self.object.project
        return context


class CatalogosView(ProyectosMenuMixin, UpdateView):
    permiso_requerido = 3
    asignacion_requerida = True
    nivel_permiso_asignado = 2

    def importar_excel(self):
        raise NotImplementedError('Este metodo debe estar en una subclase.')

    def post(self, request, *args, **kwargs):
        if 'excel-file' in request.FILES.keys():
            self.importar_excel()
        return super().post(request, *args, **kwargs)

    def get_assignment_args(self):
        self.object = self.get_object()
        return self.object, self.request.user.contrato_set.all()

    def get_object(self):
        if not hasattr(self, 'object'):
            self.model = self.form_class.fk.related_model._meta.model
            self.object = get_object_403_or_404(
                self.model,
                self.request.user,
                contraparte__company=self.request.user.currently_at,
                pk=self.kwargs['pk']
            )
        return self.object

    def get_success_url(self):
        return reverse(
            'construbot.proyectos:contrato_detail',
            kwargs={'pk': self.kwargs['pk']}
        )

    def get_context_data(self, **kwargs):
        context = super(CatalogosView, self).get_context_data(**kwargs)
        context['type'] = self.tipo if hasattr(self, 'tipo') else None
        context['formset'] = context.pop('form')
        return context


class CatalogoRetencionesInlineFormView(CatalogosView):
    change_company_ability = False
    form_class = forms.ContractRetentionInlineForm
    template_name = 'proyectos/catalogo-conceptos-inline.html'
    tipo = 'retenciones'

    def importar_excel(self):
        contrato_instance = shortcuts.get_object_or_404(
            Contrato, pk=self.request.POST['contrato'],
            contraparte__company=self.request.user.currently_at
        )
        ws = load_workbook(self.request.FILES['excel-file']).active
        for row in ws.iter_rows(min_row=2, max_col=5, max_row=ws.max_row, values_only=True):
            nombre = row[0]
            tipo = None
            if row[1].lower() == 'porcentaje':
                tipo = 'PERCENTAGE'
            elif row[1].lower() == 'monto':
                tipo = 'AMOUNT'
            valor = row[2]
            Retenciones.objects.create(
                nombre=nombre, valor=valor, tipo=tipo, project=contrato_instance
            )


class CatalogoConceptosInlineFormView(CatalogosView):
    change_company_ability = False
    form_class = forms.ContractConceptInlineForm
    template_name = 'proyectos/catalogo-conceptos-inline.html'
    tipo = 'conceptos'

    def importar_excel(self):
        contrato_instance = shortcuts.get_object_or_404(
            Contrato, pk=self.request.POST['contrato'], contraparte__company=self.request.user.currently_at)
        unidades = {}
        ws = load_workbook(self.request.FILES['excel-file']).active
        for row in ws.iter_rows(min_row=2, max_col=5, max_row=ws.max_row, values_only=True):
            codigo = row[0]
            concept_text = " ".join(row[1].splitlines())
            unidad = row[2]
            cantidad = Decimal(row[3])
            pu = row[4]
            if pu in unidades.keys():
                pu = unidades[pu]
            else:
                unidad, created = Units.objects.get_or_create(unit=unidad, company=self.request.user.currently_at)
            Concept.objects.create(
                code=codigo, concept_text=concept_text, project=contrato_instance,
                unit=unidad, total_cuantity=cantidad, unit_price=pu
            )


class CatalogoUnitsInlineFormView(CatalogosView):
    permiso_requerido = 1
    asignacion_requerida = False
    nivel_permiso_asignado = 1
    form_class = forms.UnitsInlineForm
    template_name = 'proyectos/catalogo-conceptos-inline.html'
    tipo = 'unidades'

    def importar_excel(self):
        pass

    def get_object(self):
        obj = self.request.user.currently_at
        return obj

    def get_success_url(self):
        return reverse(
            'construbot.proyectos:proyect_dashboard'
        )


class DynamicDelete(ProyectosMenuMixin, DeleteView):
    permiso_requerido = 3
    nivel_permiso_asignado = 2
    template_name = 'core/delete.html'

    def get_nivel_permiso(self):
        self.object = self.get_object()
        if isinstance(self.object, (Contrato, Estimate)):
            return self.enforce_assignment(*self.get_assignment_args())
        elif isinstance(self.object, Contraparte):
            return self.permiso_requerido
        return self.nivel_permiso_asignado

    def get_assignment_args(self):
        if isinstance(self.object, Contrato):
            return self.object, self.request.user.contrato_set.all()
        return self.object.project, self.request.user.contrato_set.all()

    def get_object(self):
        if not hasattr(self, 'object'):
            self.model = self.model_options[self.kwargs['model']]['model']
            self.object = get_object_403_or_404(
                self.model,
                self.request.user,
                **self.get_company_query(self.kwargs['model']),
                pk=self.kwargs['pk']
            )
        return self.object

    def get_company_query(self, opcion):
        kwargs = super(DynamicDelete, self).get_company_query(opcion)
        if hasattr(self, 'object'):
            if opcion == 'Estimate':
                folio = {'consecutive__gt': self.object.consecutive}
                field = 'consecutive'
            else:
                folio = {'folio__gt': self.object.folio}
                field = 'folio'
            kwargs.update(folio)
            return (kwargs, field)
        return kwargs

    def folio_handling(self):
        if hasattr(self.object, 'folio') or hasattr(self.object, 'consecutive'):
            kwargs, field = self.get_company_query(self.kwargs['model'])
            if(hasattr(self.object, 'folio')):
                self.model.objects.filter(**kwargs).update(folio=F(field) - 1)
            else:
                self.model.objects.filter(**kwargs).update(consecutive=F(field) - 1)

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.folio_handling()
        self.object.delete()
        return JsonResponse({"exito": True})


class AutocompletePoryectos(BasicAutocomplete):
    permiso_requerido = 1
    app_label_name = ProyectosConfig.verbose_name

    def get_key_words(self):
        base_string = '__unaccent__icontains'
        if self.create_field:
            search_string = self.create_field + base_string
            return {search_string: self.q}
        else:
            return {}


class ClienteAutocomplete(AutocompletePoryectos):
    model = Contraparte
    ordering = 'cliente_name'
    tipo = 'CLIENTE'

    def get_key_words(self):
        key_words = super(ClienteAutocomplete, self).get_key_words()
        key_words.update(
            {'company': self.request.user.currently_at, 'tipo': self.tipo})
        return key_words

    def get_post_key_words(self):
        kw = {'company': self.request.user.currently_at, 'tipo': self.tipo}
        return kw


class SubcontratistaAutocomplete(ClienteAutocomplete):
    tipo = 'SUBCONTRATISTA'


class DestajistaAutocomplete(ClienteAutocomplete):
    tipo = 'DESTAJISTA'


class SitioAutocomplete(AutocompletePoryectos):
    model = Sitio
    ordering = 'sitio_name'

    def get_key_words(self):
        key_words = super(SitioAutocomplete, self).get_key_words()
        key_words.update({'cliente__company': self.request.user.currently_at})
        return key_words

    def get_post_key_words(self):
        # Depende enteramente de la existencia de destinatario en el
        # formulario... suceptible a errores....
        cliente = get_object_403_or_404(Contraparte, self.request.user, pk=int(self.forwarded.get('contraparte')))
        kw = {'cliente': cliente}
        return kw


class DestinatarioAutocomplete(AutocompletePoryectos):
    model = Destinatario
    ordering = 'destinatario_text'

    def get_key_words(self):
        key_words = super(DestinatarioAutocomplete, self).get_key_words()
        key_words.update({'contraparte': Contrato.objects.get(pk=int(self.forwarded.get('project'))).contraparte})
        return key_words

    def get_post_key_words(self):
        kw = {'contraparte': Contrato.objects.get(pk=int(self.forwarded.get('project'))).contraparte}
        return kw


class UnitAutocomplete(AutocompletePoryectos):
    model = Units
    ordering = 'unit'

    def get_key_words(self):
        if hasattr(self, 'create_field') and self.create_field is not None:
            key_words = super(UnitAutocomplete, self).get_key_words()
            key_words.update({'company': self.request.user.currently_at})
        else:
            key_words = {
                'unit__unaccent__icontains': self.q,
                'company': self.request.user.currently_at
            }
        return key_words

    def get_post_key_words(self):
        kw = {'company': self.request.user.currently_at}
        return kw


class UserAutocomplete(AutocompletePoryectos):
    model = User
    ordering = 'username'

    def get_queryset(self):
        if self.request.user and self.q:
            qs = self.model.objects.filter(
                (Q(username__unaccent__icontains=self.q) | Q(email__unaccent__icontains=self.q)),
                company=self.request.user.currently_at
            )
            return qs
        elif self.request.user and self.request.POST:
            return self.model.objects


class CompanyAutocomplete(AutocompletePoryectos):
    model = Company
    ordering = 'company_name'

    def get_key_words(self):
        key_words = {
            'company_name__unaccent__icontains': self.q,
        }
        return key_words

    def get_queryset(self):
        if self.request.user and self.q:
            qs = self.request.user.company.filter(**self.get_key_words()).order_by(self.ordering)
            return qs
        elif self.request.user and self.request.POST:
            return self.model.objects


class NivelAccesoAutocomplete(AutocompletePoryectos):
    model = NivelAcceso
    ordering = 'nivel'

    def get_queryset(self):
        qs = self.model.objects.filter(**self.get_key_words()).order_by(self.ordering)
        return qs

    def get_key_words(self):
        nivel = self.request.user.nivel_acceso.nivel
        if nivel >= 3:
            key_words = {
                'nivel__lte': nivel
            }
        else:
            key_words = {
                'nivel__lt': nivel
            }
        if self.q:
            key_words['nombre__unaccent__icontains'] = self.q
        return key_words


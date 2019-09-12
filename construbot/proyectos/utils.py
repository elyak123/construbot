from decimal import Decimal
from django.db.models import Sum, F
from django import shortcuts
from openpyxl import load_workbook
from construbot.core.utils import Round
from .models import Contrato, Estimate, Concept, Units, Retenciones


def contratosvigentes(user):
    if user.nivel_acceso.nivel >= 3:
        contratos = Contrato.objects.select_related('contraparte').filter(
            status=True, contraparte__company=user.currently_at, depth=1).annotate(total_estimado=Round(Sum(
                F('estimate__estimateconcept__cuantity_estimated') *
                F('estimate__estimateconcept__concept__unit_price')
            ) / F('monto') * 100
            )).order_by('-monto')
    elif user.nivel_acceso.nivel == 2:
        contratos = Contrato.objects.select_related('contraparte').filter(
            status=True, contraparte__company=user.currently_at, users=user, depth=1).annotate(total_estimado=Round(
                Sum(F('estimate__estimateconcept__cuantity_estimated') *
                    F('estimate__estimateconcept__concept__unit_price')) / F('monto') * 100
            )).order_by('-monto')
    else:
        contratos = Contrato.objects.select_related('contraparte').filter(
            status=True, contraparte__company=user.currently_at, users=user, depth=1
            ).order_by('-folio')
    return contratos


def sumatoria_query(queryset, campo):
    return queryset.aggregate(total=Sum(campo))


def estimacionespendientes_facturacion(company, almenos_coordinador, user):
    kw = {'project__contraparte__company': company, 'invoiced': False}
    if not almenos_coordinador:
        kw['project__users'] = user
    return Estimate.objects.select_related('project').filter(**kw)


def estimacionespendientes_pago(company, almenos_coordinador, user):
    kw = {'project__contraparte__company': company, 'invoiced': True, 'paid': False}
    if not almenos_coordinador:
        kw['project__users'] = user
    return Estimate.objects.select_related('project').filter(**kw)


def total_sinpago(estimaciones):
    return estimaciones.aggregate(total=Round(Sum(F('estimateconcept__cuantity_estimated') * F('concept__unit_price'))))['total']


def totalsinfacturar(estimaciones):
    return estimaciones.aggregate(
        total=Round(Sum(F('estimateconcept__cuantity_estimated') * F('concept__unit_price')))
    )['total']


def importar_catalogo_conceptos_excel(contrato_id, excel, currently_at):
    contrato_instance = shortcuts.get_object_or_404(Contrato, pk=contrato_id, cliente__company=currently_at)
    unidades = {}
    ws = load_workbook(excel).active
    for row in ws.iter_rows(min_row=2, max_col=5, max_row=ws.max_row, values_only=True):
        codigo = row[0]
        concept_text = row[1]
        unidad = row[2]
        cantidad = Decimal(row[3])
        pu = row[4]
        if pu in unidades.keys():
            pu = unidades[pu]
        else:
            unidad, created = Units.objects.get_or_create(unit=unidad, company=currently_at)
        Concept.objects.create(
            code=codigo, concept_text=concept_text, project=contrato_instance,
            unit=unidad, total_cuantity=cantidad, unit_price=pu
        )


def importar_catalogo_retenciones_excel(contrato_id, excel, currently_at):
    import pdb; pdb.set_trace()
    contrato_instance = shortcuts.get_object_or_404(Contrato, pk=contrato_id, cliente__company=currently_at)
    unidades = {}
    ws = load_workbook(excel).active
    import pdb; pdb.set_trace()
    for row in ws.iter_rows(min_row=2, max_col=5, max_row=ws.max_row, values_only=True):
        nombre = row[0]
        tipo = ""
        if row[1].lower() == 'porcentaje':
            tipo = 'PERCENTAGE'
        elif row[1].lower() == 'monto':
            tipo = 'AMOUNT'
        valor = row[2]
        try:
            Retenciones.objects.create(
                nombre=nombre, valor=valor, tipo=tipo, project=contrato_instance
            )
        except Exception as e:
            pass

